import os
from django.http import HttpResponseRedirect, Http404
from django.shortcuts import render
from django.urls import reverse_lazy
from .models import *
from django.views import View
from django.core.files.storage import FileSystemStorage
import pandas as pd
from openpyxl import Workbook
import datetime
from django.shortcuts import redirect
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font

############################
# 1) Razom_number = 65197
#    Route = 387
# 2) Razom_number = 65198
#    Route = 209
#############################

class ToolsManager(View):
    template_name = 'ToolsManager.html'

    def get(self, request):
        if request.user.is_staff:
            content = {'user': User.objects.filter(pk=request.user.id)}
            return render(request, self.template_name, content)
        return HttpResponseRedirect(reverse_lazy('Autentificate'))

    def post(self, request):
        if request.user.is_staff:
            if request.FILES['myfile']:
                myfile = request.FILES['myfile']
                fs = FileSystemStorage()
                filename = fs.save(myfile.name, myfile)
                excel_file = fs.url(filename)
                df = pd.read_excel('.'+ excel_file).to_numpy()
                Rk_list = [i[0] for i in df]
                li = Totalplanes.objects.filter(Razom_number__in=Rk_list).values('route')
                print(Rk_list)

                # set configuration
                wb = Workbook()
                wbAdmin = wb.active

                # set column width
                wbAdmin.title = 'Изменения Угла'
                wbAdmin.column_dimensions['A'].width = 9
                wbAdmin.column_dimensions['B'].width = 8
                wbAdmin.column_dimensions['C'].width = 11
                wbAdmin.column_dimensions['D'].width = 3.5
                wbAdmin.column_dimensions['E'].width = 11

                wbAdmin['A1'] = datetime.datetime.now().strftime("%d-%m-%Y")
                wbAdmin['B3'] = 'RN'
                wbAdmin['C3'] = 'Last Version'
                wbAdmin['C3'].fill = PatternFill('solid', start_color='00FF9900')

                wbAdmin['E3'] = 'New Version'
                wbAdmin['E3'].fill = PatternFill('solid', start_color='5cb800')


                RN_start, new_start, last_start = 4, 4, 4

                for count, i in enumerate(zip(df, li)):
                    wbAdmin[f'B{RN_start+count}'] = i[0][0]
                    wbAdmin[f'E{new_start + count}'] = i[0][1]
                    wbAdmin[f'D{new_start + count}'] = ">>>"
                    wbAdmin[f'D{new_start + count}'].font = Font(color="0099CC00")

                    wbAdmin[f'C{last_start + count}'] = i[1]['route']

                # PRODUCTION ROUTE
                # strPath = "/home/acrzmcomua/public_html/"
                # filename = strPath + f"media/OhhRazom/Excel/ManagerToolsRoutes/change_route_{datetime.datetime.now().strftime('%d_%m_%Y')}.xlsx"

                # DEPLOYMENT ROUTE
                filename = f"media\OhhRazom\Excel\ManagerToolsRoutes\change_route_{datetime.datetime.now().strftime('%d_%m_%Y')}.xlsx"
                wb.save(filename=filename)

                for i in df:
                    Totalplanes.objects.filter(Razom_number=i[0]).update(route=i[1])

                # PRODUCTUIN ROUTE
                # return redirect(f"/media/OhhRazom/Excel/ManagerToolsRoutes/change_route_{datetime.datetime.now().strftime('%d_%m_%Y')}.xlsx")

                # DEPLOYMENT ROUTE
                return redirect(f"/media/OhhRazom/Excel/ManagerToolsRoutes/change_route_{datetime.datetime.now().strftime('%d_%m_%Y')}.xlsx")
        return HttpResponseRedirect(reverse_lazy('Autentificate'))


class ToolsManagerBot(View):
    template_name = 'ToolsManagerBot.html'

    def get(self, request):
        if request.user.is_staff:
            return render(request, self.template_name)
        return HttpResponseRedirect(reverse_lazy('Autentificate'))


