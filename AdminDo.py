import os

from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views import View

from .AppFiles.LengualeOut import WorkFile
from .models import *

import datetime
import traceback
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment


class WorkAdminPage(View):

    def post(self, request, flag):
        if flag == "ChangeLenguage":
            request.session['Currentpage'] = "Admin"
            request.session['inLenluage'] = request.POST.get('Len')
            return HttpResponseRedirect(reverse_lazy('ChangeLenguage'))
        elif flag == "DeleteMarkMap":
            pass
        else:
            raise Http404


class AsyincDelete(View):

    def post(self, request, action):

        if action == "DeleteMarkMap":
            try:
                instanceRK = RkCompany.objects.filter(rk=request.session['currentOutAC']).get(
                    Razom_number__Razom_number=int(request.POST.get('numb')))
                instanceRK.delete()
                return HttpResponse(f"Mark Delete {request.POST.get('numb')}")
            except Exception as e:
                print(traceback.format_exc())
                return HttpResponse(f"Don`t mark Delete {request.POST.get('numb')}")
        else:
            raise Http404


class ExportExcelAdmin(View):

    def post(self, request):
        try:

            dictHead = WorkFile.outLenguage("tableHead")
            leterPostion = {
                "A": dictHead['Number'],
                "B": dictHead['city'],
                "C": dictHead['adress'],
                "D": dictHead['type'],
                "E": dictHead['format'],
                "F": dictHead['side'],
                "G": {
                    "RU": "OTS",
                    "EN": "OTS",
                    "UA": "OTS"
                },
                "H": {
                    "RU": "GRP",
                    "EN": "GRP",
                    "UA": "GRP"
                },
                "I": {
                    "RU": "DOORS",
                    "EN": "DOORS",
                    "UA": "DOORS"
                },
                "J": dictHead['SelfNumberCont'],
                "K": dictHead['Razom_number'],
                "L": {
                    "RU": "????????",
                    "EN": "Photo",
                    "UA": "????????"
                },
                "M": {
                    "RU": "??????????",
                    "EN": "Scheme",
                    "UA": "??????????"
                },
                "N": dictHead['story'],
                "O": dictHead['contactor'],
            }
            wb = Workbook()
            NameRK = Rk.objects.get(id=request.session['currentOutAC'])
            #strPath = ""
            strPath = "/home/acrzmcomua/public_html/"
            # strPath = "D:\work\Razom\ooh\\" #?????? ????????
            filename = strPath + f"media/OhhRazom/Excel/{request.user}{NameRK.RK}.xlsx"
            wbAdmin = wb.active
            wbAdmin.title = "AC"
            stratCursor = '4'
            thin = Side(border_style="thin", color="081810")
            wbAdmin.row_dimensions[4].height = 25
            wbAdmin.column_dimensions['A'].width = 14
            wbAdmin.column_dimensions['B'].width = 16
            wbAdmin.column_dimensions['C'].width = 55
            wbAdmin.column_dimensions['D'].width = 16
            wbAdmin.column_dimensions['E'].width = 16
            wbAdmin.column_dimensions['F'].width = 8
            wbAdmin.column_dimensions['G'].width = 16
            wbAdmin.column_dimensions['H'].width = 10
            wbAdmin.column_dimensions['I'].width = 18
            wbAdmin.column_dimensions['J'].width = 25
            wbAdmin.column_dimensions['K'].width = 25
            wbAdmin.column_dimensions['L'].width = 16
            wbAdmin.column_dimensions['M'].width = 16
            wbAdmin.column_dimensions['N'].width = 16
            wbAdmin.column_dimensions['O'].width = 20

            if request.COOKIES['lenguage'] == "RU":
                wbAdmin["A1"] = "???????? ????????????????: "
                wbAdmin["A2"] = "??????????????: "
            elif request.COOKIES['lenguage'] == "EN":
                wbAdmin["A1"] = "Date of export: "
                wbAdmin["A2"] = "Product: "
            elif request.COOKIES['lenguage'] == "UA":
                wbAdmin["A1"] = "???????? ????????????????: "
                wbAdmin["A2"] = "??????????????: "
            else:
                wbAdmin["A1"] = "???????? ????????????????: "
                wbAdmin["A2"] = "??????????????: "

            now = datetime.datetime.now()

            wbAdmin["B1"] = now.strftime("%d-%m-%Y")
            wbAdmin["B2"] = NameRK.RK

            for i in leterPostion:
                wbAdmin[f"{i}{stratCursor}"].fill = PatternFill("solid", "3b96f7")
                wbAdmin[f"{i}{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                if request.COOKIES['lenguage'] == "RU":
                    wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["RU"]
                elif request.COOKIES['lenguage'] == "EN":
                    wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["EN"]
                elif request.COOKIES['lenguage'] == "UA":
                    wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["UA"]
                else:
                    wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["RU"]

            stratCursor = 5
            dataExport = request.session['currentDataOut']
            for count, j in enumerate(dataExport):
                wbAdmin[f"A{stratCursor}"] = count + 1
                tmpInst = RkCompany.objects.filter(rk=int(request.session["currentOutAC"])).get(
                    Razom_number=j['Razom_number_id'])
                for k in leterPostion:
                    wbAdmin[f"{k}{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[f"{k}{stratCursor}"].alignment = Alignment(horizontal='left')
                if tmpInst.Grade == 1:
                    for k in leterPostion:
                        wbAdmin[f'{k}{stratCursor}'].fill = PatternFill("solid", "32CD32")
                elif tmpInst.Grade == 2:
                    for k in leterPostion:
                        wbAdmin[f'{k}{stratCursor}'].fill = PatternFill("solid", "FFA07A")
                if request.COOKIES['lenguage'] == "RU":
                    wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_RU
                    wbAdmin[f"C{stratCursor}"] = f'{tmpInst.Razom_number.adress.adress_RU} {tmpInst.Razom_number.house} {tmpInst.Razom_number.loc.Location_RU}'
                    wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeRU
                    wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                    wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                    wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                    wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                    wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                    wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                    wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                    wbAdmin[
                        f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                    wbAdmin[f"L{stratCursor}"] = "????????"
                    wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[
                        f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                    wbAdmin[f"M{stratCursor}"] = "??????????"
                    wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                    wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_RU
                elif request.COOKIES['lenguage'] == "EN":
                    wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_EN
                    wbAdmin[f"C{stratCursor}"] = f'{tmpInst.Razom_number.adress.adress_EN} {tmpInst.Razom_number.house} {tmpInst.Razom_number.loc.Location_EN}'
                    wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeEN
                    wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                    wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                    wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                    wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                    wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                    wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                    wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                    wbAdmin[
                        f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                    wbAdmin[f"L{stratCursor}"] = "Photo"
                    wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[
                        f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                    wbAdmin[f"M{stratCursor}"] = "Scheme"
                    wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                    wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_EN
                elif request.COOKIES['lenguage'] == "UA":
                    wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_UA
                    wbAdmin[f"C{stratCursor}"] = f'{tmpInst.Razom_number.adress.adress_UA} {tmpInst.Razom_number.house} {tmpInst.Razom_number.loc.Location_UA}'
                    wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeUA
                    wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                    wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                    wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                    wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                    wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                    wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                    wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                    wbAdmin[
                        f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                    wbAdmin[f"L{stratCursor}"] = "????????"
                    wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[
                        f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                    wbAdmin[f"M{stratCursor}"] = "??????????"
                    wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                    wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_UA
                else:
                    wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_RU
                    wbAdmin[f"C{stratCursor}"] = tmpInst.Razom_number.adress.adress_RU
                    wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeRU
                    wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                    wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                    wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                    wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                    wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                    wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                    wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                    wbAdmin[
                        f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                    wbAdmin[f"L{stratCursor}"] = "????????"
                    wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[
                        f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                    wbAdmin[f"M{stratCursor}"] = "??????????"
                    wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                    wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                    wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_RU
                wbAdmin[f"N{stratCursor}"].fill = PatternFill("solid", str(tmpInst.story.color).replace("#", ""))
                stratCursor += 1
            # strPath = "D:\?????????????? ???????? ??\ooh" #windows
            strPath = "/home/acrzmcomua/public_html"  # linux
            # strPath = 'D:\work\Razom\ooh' #?????? ????????
            if os.path.exists(os.path.join(strPath, "media", "OhhRazom", "Excel", f'{request.user}{NameRK.RK}.xlsx')):
                os.remove(os.path.join(strPath, "media", "OhhRazom", "Excel", f'{request.user}{NameRK.RK}.xlsx'))
            wb.save(filename=filename)
            # return redirect(f'/media/OhhRazom/Excel/{request.user}{NameRK.RK}.xlsx') # ?????? ????????
            return redirect(f'/media/OhhRazom/Excel/{request.user}{NameRK.RK}.xlsx')
        except Exception as e:
            print(traceback.format_exc())
            return HttpResponse(traceback.format_exc())
