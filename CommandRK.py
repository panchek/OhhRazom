import os

from django.http import HttpResponseRedirect, Http404, HttpResponse
from django.shortcuts import redirect
from django.urls import reverse_lazy
from django.views import View
from openpyxl import Workbook

from .AppFiles.LengualeOut import WorkFile
from .models import *
from django.contrib import messages
import traceback
from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import datetime


class OutputWindow:

    @staticmethod
    def numberIn(text):
        outList = []
        for i in (str(text).replace("\n", " ")).split():
            outList.append(i.strip())
        return outList

class CommandWork(View):

    def post(self, request, utils):
        if utils == "AddMark":
            errorNumb = "Ошибка с такими кодами : "
            try:
                if request.POST.get("story") == "":
                    story = "Main visual"
                else:
                    story = request.POST.get("story")
                try:
                    infStory = Story.objects.filter(rk=request.session['SettingRKName']).get(story=story)
                except:
                    try:
                        if request.POST.get("color_begin") == "":
                            color = "#191970"
                        else:
                            color = request.POST.get("color_begin")
                        infStory = Story(
                            rk= Rk.objects.get(id=request.session['SettingRKName']),
                            story= story,
                            color= color
                        )
                        infStory.save()
                    except Exception as e:
                        print('Ошибка:\n', traceback.format_exc())
                if request.POST.get("choise_type") == "doors":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        print(i)
                        try:
                            RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__doors=i)
                            errorNumb += str(i) + " "
                        except:
                            try:
                                tmpNumb = Totalplanes.objects.get(doors=i)
                                RkCompany(
                                    rk= Rk.objects.get(id=request.session['SettingRKName']),
                                    Razom_number= tmpNumb,
                                    story= infStory
                                ).save()
                            except Exception as e:
                                print('Ошибка:\n', traceback.format_exc())
                                print(f"Нет такого {i}")
                                errorNumb += str(i) + " "

                elif request.POST.get("choise_type") == "Razom_number":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        try:
                            RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__Razom_number=int(i))
                            errorNumb += str(i) + " "
                        except:
                            try:
                                tmpNumb = Totalplanes.objects.get(Razom_number=int(i))
                                RkCompany(
                                    rk=Rk.objects.get(id=request.session['SettingRKName']),
                                    Razom_number=tmpNumb,
                                    story=infStory
                                ).save()
                            except Exception as e:
                                print('Ошибка:\n', traceback.format_exc())
                                print(f"Нет такого {i}")
                                errorNumb += str(i) + " "
                else:
                    print("Not Work")
                if errorNumb != "":
                    print("hello")
                    messages.success(request, errorNumb)
            except Exception as e:
                print('Ошибка:\n', traceback.format_exc())
                messages.success(request, 'Ошибка')
            finally:
                return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        elif utils == "Changecolor":
            errorNumb = "Ошибка с такими кодами : "
            try:
                if request.POST.get("story") == "":
                    story = "Main visual"
                else:
                    story = request.POST.get("story")
                try:
                    infStory = Story.objects.filter(rk=request.session['SettingRKName']).get(story=story)
                except:
                    try:
                        if request.POST.get("color_begin") == "":
                            color = "#191970"
                        else:
                            color = request.POST.get("color_begin")
                        infStory = Story(
                            rk=Rk.objects.get(id=request.session['SettingRKName']),
                            story=story,
                            color=color
                        )
                        infStory.save()
                    except Exception as e:
                        print('Ошибка:\n', traceback.format_exc())
                if request.POST.get("choise_type") == "doors":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        try:
                            tmpRKNumb = RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__doors=i)
                            tmpRKNumb.story = infStory
                            tmpRKNumb.save()
                        except Exception as e:
                            print('Ошибка:\n', traceback.format_exc())
                            print(f"Нет такого {i}")
                            errorNumb = str(i) + " ы"
                elif request.POST.get("choise_type") == "Razom_number":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        try:
                            tmpRKNumb = RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__Razom_number=int(i))
                            tmpRKNumb.story = infStory
                            tmpRKNumb.save()
                        except Exception as e:
                            print('Ошибка:\n', traceback.format_exc())
                            print(f"Нет такого {i}")
                            errorNumb = str(i) + " "
                else:
                    print("Not Work")
                if errorNumb != "":
                    messages.success(request, errorNumb)
            except Exception as e:
                print('Ошибка:\n', traceback.format_exc())
                messages.success(request, 'Ошибка')
            finally:
                return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        elif utils == "DeleteMark":
            try:
                if request.POST.get("choise_type") == "doors":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        try:
                            tmpRKNumb = RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__doors=i)
                            tmpRKNumb.delete()
                        except Exception as e:
                            print('Ошибка:\n', traceback.format_exc())
                            print(f"Нет такого {i}")
                            errorNumb = str(i) + " "
                elif request.POST.get("choise_type") == "Razom_number":
                    for i in OutputWindow.numberIn(request.POST.get("area_mark")):
                        try:
                            tmpRKNumb = RkCompany.objects.filter(rk=request.session['SettingRKName']).get(Razom_number__Razom_number=int(i))
                            tmpRKNumb.delete()
                        except Exception as e:
                            print('Ошибка:\n', traceback.format_exc())
                            print(f"Нет такого {i}")
                            errorNumb = str(i) + " "
                else:
                    print("Not Work")
                if errorNumb != "":
                    messages.success(request, errorNumb)
            except Exception as e:
                print('Ошибка:\n', traceback.format_exc())
                messages.success(request, 'Ошибка')
            finally:
                return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        elif utils == "ChangeLenguage":
            request.session['Currentpage'] = "CurrentRK"
            request.session['inLenluage'] = request.POST.get('Len')
            return HttpResponseRedirect(reverse_lazy('ChangeLenguage'))
        elif utils == "ShowMap":
            request.session['currentOutAC'] = request.session['SettingRKName']
            del request.session['SettingRKName']
            return HttpResponseRedirect(reverse_lazy('Admin'))
        elif utils == "ResetColorStory":
            try:
                instanceStory = Story.objects.filter(rk=request.session['SettingRKName']).get(story='Main visual')
            except:
                instanceStory = Story(
                    rk=Rk.objects.get(id=request.session['SettingRKName']),
                    story='Main visual',
                    color='#191970'
                )
                instanceStory.save()
            allObjsectInAC = RkCompany.objects.filter(rk=request.session['SettingRKName'])
            for i in allObjsectInAC:
                i.story = instanceStory
                i.save()
            return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        elif utils == "ClearAC":
            instanceAC = RkCompany.objects.filter(rk=request.session['SettingRKName'])
            for i in instanceAC:
                i.delete()
            return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        elif utils == "back":
            return HttpResponseRedirect(reverse_lazy('setting'))
        elif utils == "ToExcel":
            try:
                dictHead = WorkFile.outLenguage("tableHead")
                leterPostion = {
                    "A": dictHead['Number'],
                    "B": dictHead['city'],
                    "C": dictHead['adress'],
                    "D": dictHead['type'],
                    "E": dictHead['format'],
                    "F": dictHead['side'],
                    "G": {
                        "RU": "OTS",
                        "EN": "OTS",
                        "UA": "OTS"
                    },
                    "H": {
                        "RU": "GRP",
                        "EN": "GRP",
                        "UA": "GRP"
                    },
                    "I": {
                        "RU": "Doors",
                        "EN": "Doors",
                        "UA": "Doors"
                    },
                    "J": dictHead['SelfNumberCont'],
                    "K": dictHead['Razom_number'],
                    "L": {
                        "RU": "Фото",
                        "EN": "Photo",
                        "UA": "Фото"
                    },
                    "M": {
                        "RU": "Схема",
                        "EN": "Scheme",
                        "UA": "Схема"
                    },
                    "N": dictHead['story'],
                    "O": dictHead['contactor'],
                }
                wb = Workbook()
                NameRK = Rk.objects.get(id=request.session['SettingRKName'])
                strPath = ""  # windows
                strPath = "/home/acrzmcomua/public_html/"  # linux
                filename = strPath + f"media/OhhRazom/Excel/{request.user}{NameRK.RK}.xlsx"
                wbAdmin = wb.active
                wbAdmin.title = "AC"
                stratCursor = '4'
                thin = Side(border_style="thin", color="081810")
                wbAdmin.row_dimensions[4].height = 25
                wbAdmin.column_dimensions['A'].width = 14
                wbAdmin.column_dimensions['B'].width = 16
                wbAdmin.column_dimensions['C'].width = 55
                wbAdmin.column_dimensions['D'].width = 16
                wbAdmin.column_dimensions['E'].width = 16
                wbAdmin.column_dimensions['F'].width = 8
                wbAdmin.column_dimensions['G'].width = 16
                wbAdmin.column_dimensions['H'].width = 10
                wbAdmin.column_dimensions['I'].width = 18
                wbAdmin.column_dimensions['J'].width = 25
                wbAdmin.column_dimensions['K'].width = 25
                wbAdmin.column_dimensions['L'].width = 16
                wbAdmin.column_dimensions['M'].width = 16
                wbAdmin.column_dimensions['N'].width = 16
                wbAdmin.column_dimensions['O'].width = 20

                if request.COOKIES['lenguage'] == "RU":
                    wbAdmin["A1"] = "Дата экспорта: "
                    wbAdmin["A2"] = "Продукт: "
                elif request.COOKIES['lenguage'] == "EN":
                    wbAdmin["A1"] = "Date of export: "
                    wbAdmin["A2"] = "Product: "
                elif request.COOKIES['lenguage'] == "UA":
                    wbAdmin["A1"] = "Дата експорту: "
                    wbAdmin["A2"] = "Продукт: "
                else:
                    wbAdmin["A1"] = "Дата экспорта: "
                    wbAdmin["A2"] = "Продукт: "

                now = datetime.datetime.now()

                wbAdmin["B1"] = now.strftime("%d-%m-%Y")
                wbAdmin["B2"] = NameRK.RK

                for i in leterPostion:
                    wbAdmin[f"{i}{stratCursor}"].fill = PatternFill("solid", "3b96f7")
                    wbAdmin[f"{i}{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    if request.COOKIES['lenguage'] == "RU":
                        wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["RU"]
                    elif request.COOKIES['lenguage'] == "EN":
                        wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["EN"]
                    elif request.COOKIES['lenguage'] == "UA":
                        wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["UA"]
                    else:
                        wbAdmin[f"{i}{stratCursor}"] = leterPostion[i]["RU"]

                stratCursor = 5
                dataExport = request.session['currentDataOut']
                for count, j in enumerate(dataExport):
                    wbAdmin[f"A{stratCursor}"] = count + 1
                    tmpInst = RkCompany.objects.filter(rk=int(request.session["SettingRKName"])).get(
                        Razom_number=j['Razom_number_id'])
                    for k in leterPostion:
                        wbAdmin[f"{k}{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[f"{k}{stratCursor}"].alignment = Alignment(horizontal='left')
                    if tmpInst.Grade == 1:
                        for k in leterPostion:
                            wbAdmin[f'{k}{stratCursor}'].fill = PatternFill("solid", "32CD32")
                    elif tmpInst.Grade == 2:
                        for k in leterPostion:
                            wbAdmin[f'{k}{stratCursor}'].fill = PatternFill("solid", "FFA07A")
                    if request.COOKIES['lenguage'] == "RU":
                        wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_RU
                        wbAdmin[f"C{stratCursor}"] = tmpInst.Razom_number.adress.adress_RU
                        wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeRU
                        wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                        wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                        wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                        wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                        wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                        wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                        wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                        wbAdmin[
                            f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                        wbAdmin[f"L{stratCursor}"] = "Фото"
                        wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[
                            f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                        wbAdmin[f"M{stratCursor}"] = "Схема"
                        wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                        wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_RU
                    elif request.COOKIES['lenguage'] == "EN":
                        wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_EN
                        wbAdmin[f"C{stratCursor}"] = tmpInst.Razom_number.adress.adress_EN
                        wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeEN
                        wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                        wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                        wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                        wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                        wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                        wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                        wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                        wbAdmin[
                            f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                        wbAdmin[f"L{stratCursor}"] = "Photo"
                        wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[
                            f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                        wbAdmin[f"M{stratCursor}"] = "Scheme"
                        wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                        wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_EN
                    elif request.COOKIES['lenguage'] == "UA":
                        wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_UA
                        wbAdmin[f"C{stratCursor}"] = tmpInst.Razom_number.adress.adress_UA
                        wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeUA
                        wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                        wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                        wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                        wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                        wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                        wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                        wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                        wbAdmin[
                            f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                        wbAdmin[f"L{stratCursor}"] = "Фото"
                        wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[
                            f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                        wbAdmin[f"M{stratCursor}"] = "Схема"
                        wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                        wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_UA
                    else:
                        wbAdmin[f"B{stratCursor}"] = tmpInst.Razom_number.city_standart.city_standart_RU
                        wbAdmin[f"C{stratCursor}"] = tmpInst.Razom_number.adress.adress_RU
                        wbAdmin[f"D{stratCursor}"] = tmpInst.Razom_number.type.typeRU
                        wbAdmin[f"E{stratCursor}"] = tmpInst.Razom_number.format.format
                        wbAdmin[f"F{stratCursor}"] = tmpInst.Razom_number.side.side
                        wbAdmin[f"G{stratCursor}"] = tmpInst.Razom_number.OTS
                        wbAdmin[f"H{stratCursor}"] = tmpInst.Razom_number.GRP
                        wbAdmin[f"I{stratCursor}"] = tmpInst.Razom_number.doors
                        wbAdmin[f"J{stratCursor}"] = tmpInst.Razom_number.Self_number
                        wbAdmin[f"K{stratCursor}"] = tmpInst.Razom_number.Razom_number
                        wbAdmin[
                            f"L{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imagePhoto}"
                        wbAdmin[f"L{stratCursor}"] = "Фото"
                        wbAdmin[f"L{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"L{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[
                            f"M{stratCursor}"].hyperlink = f"https://ac.rzm.com.ua/media/{tmpInst.Razom_number.imageShema}"
                        wbAdmin[f"M{stratCursor}"] = "Схема"
                        wbAdmin[f"M{stratCursor}"].style = "Hyperlink"
                        wbAdmin[f"M{stratCursor}"].border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        wbAdmin[f"N{stratCursor}"] = tmpInst.story.story
                        wbAdmin[f"O{stratCursor}"] = tmpInst.Razom_number.Contractor.Contractor_RU
                    wbAdmin[f"N{stratCursor}"].fill = PatternFill("solid", str(tmpInst.story.color).replace("#", ""))
                    stratCursor += 1
                # strPath = "D:\перевод сайт в\ooh" #windows
                strPath = "/home/acrzmcomua/public_html"  # linux
                if os.path.exists(
                        os.path.join(strPath, "media", "OhhRazom", "Excel", f'{request.user}{NameRK.RK}.xlsx')):
                    os.remove(os.path.join(strPath, "media", "OhhRazom", "Excel", f'{request.user}{NameRK.RK}.xlsx'))
                wb.save(filename=filename)
                return redirect(f'/media/OhhRazom/Excel/{request.user}{NameRK.RK}.xlsx')
            except Exception as e:
                print(traceback.format_exc())
                return HttpResponse(traceback.format_exc())
        elif utils == "AddReach":
            try :
                inst = ReachCity.objects.filter(rk=int(request.session['SettingRKName'])).get(city=int(request.POST.get('ChooseCity')))
                inst.reach=  request.POST.get('TextForReach')
                inst.frequency =  request.POST.get('TextForFrequency')
                inst.period =  request.POST.get('TextForPeriod')
                inst.save()
            except:
                Instance = ReachCity(
                    rk = Rk.objects.get(id=request.session['SettingRKName']),
                    city = CityPlanes.objects.get(id=int(request.POST.get('ChooseCity'))),
                    reach = request.POST.get('TextForReach'),
                    frequency=request.POST.get('TextForFrequency'),
                    period=request.POST.get('TextForPeriod')
                )
                Instance.save()
            return HttpResponseRedirect(reverse_lazy('CurrentRK'))
        else:
            raise Http404
